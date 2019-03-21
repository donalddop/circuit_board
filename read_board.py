''' Author: Donald Nikkessen

This file contains functions to read a board specification from an xlsx file
into a python usable form.

The external library openpyxl is used to read the xlsx file.
'''

from openpyxl import load_workbook

def process_file(filename):
    ''' Process an xlsx file containing gate and connection data.

    Inputs:
        filename: name of the file to be processed
    Returns:
        gate_dict: a dict mapping gate numbers to their coordinates.
        netlists: a list of lists each containing a set of connections.
    '''
    # Initializing variables and datastructures
    read_gate = False
    read_netlist = False
    gate_dict = {}
    netlists = []
    # Open a workbook from the file
    wb = load_workbook(filename)
    ws = wb.active
    print('Reading file for: ', ws['A1'].value, ws['B1'].value)
    for row in ws.rows:
        if read_gate == True:
            # When we are done with the gates we start reading netlists
            if row[0].value == None:
                read_gate = False
                continue
            # For each gate number create a dict entry {gate : (0,x,y)}
            gate_dict[row[0].value] = (0,row[1].value,row[2].value)
        # Read all connections from a netlist into a list of tuples
        if read_netlist == True:
            if row[0].value == None or row == list(ws.rows)[ws.max_row - 1]:
                read_netlist = False
                netlists.append(connections)
                continue
            connections.append( (row[0].value, row[1].value) )
        # Check for next gate list or netlist to read
        if row[0].value == 'Gate number':
            read_gate = True
        if row[0].value == 'First gate number':
            read_netlist = True
            connections = []

    return gate_dict, netlists
